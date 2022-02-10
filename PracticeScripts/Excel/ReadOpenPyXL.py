import openpyxl
path = "gfg.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
row = sheet_obj.max_row
column = sheet_obj.max_column
print("total Rows: ",row)
print("total columns: ",column)

print("\nValue of first column")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)

print("\nValue of first row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")

cell_obj = sheet_obj.cell(row = 1,column = 1)
print(cell_obj.value)