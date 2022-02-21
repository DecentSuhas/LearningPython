import random

import openpyxl

from PracticeScripts.StudentManagement.AdminLogin import AdminLogin
from PracticeScripts.StudentManagement.DB_Operations import ConnectToDB
import pandas as pd
from matplotlib import pyplot as plt


def export_students_details(tablename, u_name):
    """
    To fetch all the records from DB and write it into excel file.
    :param tablename:
    :return:
    """
    db = ConnectToDB()
    fileName = "Students_Details.xlsx"
    try:
        admin = AdminLogin()
        if admin.verify_login("admin_credentials", u_name):
            dbs1 = ConnectToDB.cursors.execute("select * from " + tablename)
            value = ConnectToDB.cursors.fetchall()
            df = pd.DataFrame(value)
            df.columns = ['First name', 'Last name', 'Email', 'Username', 'Password']
            writer = pd.ExcelWriter(fileName, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            print("Details exported to .xlsx file successfully")
        else:
            print("Failed to export")
    except Exception as e:
        print(e)
        ConnectToDB.myconnection.rollback()


def export_students_records(tablename, u_name):
    """
    To fetch all the records from DB and write it into excel file.
    :param tablename:
    :return:
    """
    db = ConnectToDB()
    fileName = "Students_Record.xlsx"
    try:
        admin = AdminLogin()
        if admin.verify_login("admin_credentials", u_name):
            dbs1 = ConnectToDB.cursors.execute("select * from " + tablename)
            value = ConnectToDB.cursors.fetchall()
            df = pd.DataFrame(value)
            df.columns = ['Name', 'Class', 'Section', 'ID']
            writer = pd.ExcelWriter(fileName, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False)
            writer.save()
            print("Details exported to .xlsx file successfully")
        else:
            print("Failed to export")
    except Exception as e:
        print(e)
        ConnectToDB.myconnection.rollback()


def generate_barGraph_StudentsCount():
    """
     This function generates the bar graph as per the number of students in each class
    :return:
    """
    class_9 = filter_pandas("Class", 9, "filter.Class")
    class_10 = filter_pandas("Class", 10, "filter.Class")
    class_11 = filter_pandas("Class", 11, "filter.Class")
    class_12 = filter_pandas("Class", 12, "filter.Class")
    data = {'9th': class_9, '10th': class_10, '11th': class_11, '12th': class_12}
    classrooms = list(data.keys())
    values = list(data.values())
    fig = plt.figure(figsize=(10, 5))
    plots = plt.bar(classrooms, values, color='green', width=0.3)
    plt.xlabel("Class")
    plt.ylabel("No. of students admitted")
    plt.title("Students count as per class")
    plt.plot(kind="bar")
    plt.show()


def filter_pandas(columnName, classRoom, column):
    """
        This method refers the current .xlsx file and filters the data based on the class room and returns the total count
    :param columnName:
    :param classRoom:
    :param column:
    :return: total count of records/rows
    """
    df = pd.read_excel("Students_Record.xlsx")
    filter = df[df[columnName] == classRoom]  # Filter based on the column name
    value = eval(
        column)  # Get only the column "Class" values also the passed string is converted to object so that dataframe accepts
    row_count = value.shape[0]  # Got the row count
    return row_count


def bulkImport_incomplete():
    df = pd.read_excel("C:\\Users\\320052425\\Downloads\\Bulk_upload.xlsx")
    name = (df['Name']).to_string(index=False)  # To get only the value without index
    class_room = (df['Class']).to_string(index=False)
    section = (df['Section']).to_string(index=False)
    id = (df['ID']).to_string(index=False)
    for i in df.iterrows():
        print(name)
        print()


def getListOfValues(columnName, expectedColumn):
    """
        To get the list of values and store it in a list.
    :param columnName:
    :param expectedColumn:
    :return: list
    """
    wb = openpyxl.load_workbook("C:\\Users\\320052425\\Downloads\\Bulk_upload.xlsx")
    sheet = wb.active
    row_count = sheet.max_row
    col_count = sheet.max_column
    list_values = []
    for i in range(1, col_count + 1):
        cell1 = sheet.cell(row=1, column=i)
        if cell1.value == expectedColumn:
            for k in range(2, row_count + 1):
                cell2 = sheet.cell(row=k, column=i)
                list_values.append(cell2.value)
    return list_values


def import_by_row(table_name):
    #db = ConnectToDB()
    df = pd.read_excel("C:\\Users\\320052425\\Downloads\\Bulk_upload.xlsx")
    row_count = len(df.index)
    check = False
    pass_count = 0
    fail_count = 0
    for i in range(0,row_count):
        current_row = df.loc[i]
        sampleList = list(current_row)
        s_name = sampleList[0]
        classroom = str(sampleList[1])
        section = sampleList[2]
        id = str(sampleList[3])
        sql = "insert into " + table_name + "(name, class, section, id) values (%s, %s, %s, %s)"
        val = (s_name, classroom, section, id)
        try:
            dbs1 = ConnectToDB.cursors.execute(sql, val)
            ConnectToDB.myconnection.commit()
            check = True
            pass_count = pass_count+1
        except Exception as e:
            print(e)
            ConnectToDB.myconnection.rollback()
            check = False
            fail_count = fail_count+1
        sampleList.clear()

    print(f"Uploaded",{pass_count},"records successfully")
    print(f"Failed to upload",{fail_count},"record(s)")

#export_students_records("student_records","admin")
#generate_barGraph_StudentsCount()
#import_by_row("student_records")
