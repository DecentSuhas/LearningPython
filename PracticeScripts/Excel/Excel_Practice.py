import re
from datetime import datetime
import datetime
import openpyxl

path = "employee.xlsx"


def createExcelFile():
    wb = openpyxl.Workbook()
    wb.save("PracticeExcel_1.xlsx")


def writeIntoFile():
    wb = openpyxl.Workbook()
    sheet = wb.active
    col1 = sheet.cell(row=1, column=1)
    col1.value = "My first value"
    wb.save("PracticeExcel_1.xlsx")


def writeIntoFile_1():
    """ To write into first row """
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, 10):
        col = sheet.cell(row=1, column=i)
        col.value = i
    wb.save("PracticeExcel_1.xlsx")


def writeIntoFile_2():
    """ To write into first column"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, 10):
        rows = sheet.cell(row=i, column=1)
        rows.value = i
    wb.save("PracticeExcel_1.xlsx")


def writeIntoFile_3():
    """ To write into first 10 rows and first 2 columns"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    for i in range(1, 10):
        rows = sheet.cell(row=i, column=1)
        col = sheet.cell(row=i, column=2)
        rows.value = "Rows"
        col.value = "Cols"
    wb.save("PracticeExcel_1.xlsx")


def readFromFile_1():
    """ Read a value from a cell"""
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    cell = sheet.cell(row=4, column=10)
    print(cell.value)


def readFromFile_2():
    """ Read value from all cells"""
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    for i in range(1, rows + 1):
        cell1 = sheet.cell(row=i, column=1)
        cell2 = sheet.cell(row=i, column=2)
        print(cell1.value, cell2.value)


def getEValuesOfGivenColumn(column_value):
    """
    Get only the values of all the rows of first column where the given column should match
    :param cell_value:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=i)
        if cell.value == column_value:
            for k in range(i, sheet.max_row + 1):
                item = sheet.cell(row=k, column=i)
                print(item.value)


def getDetailsOfaRecord(record, column_name):
    """
    Get the complete details of the given record(name)
    :param record, column_name:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column

    for i in range(1, cols + 1):
        column_value = sheet.cell(row=1, column=i)
        if column_value.value == column_name:
            for k in range(i, rows + 1):
                cell = sheet.cell(row=k, column=i)
                if cell.value == record:
                    for j in range(1, cols + 1):
                        getValue = sheet.cell(row=k, column=j)
                        print(getValue.value)


def findRecordInWholeFile(record):  # Not complete
    """
    To find the presence of the value in the whole file
    :param record:
    :return:
    """

    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_names
    # print(sheet.)


def getAllDuplicateRecords(record):
    """
    Write a program to print all records of duplicates where first name is same
    :param record: Pass the first name
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    for i in range(1, row_count + 1):
        cell_1 = sheet.cell(row=i, column=2)
        if cell_1.value == record:
            for k in range(1, column_count + 1):
                cell_2 = sheet.cell(row=i, column=k)
                print(cell_2.value)



def getValueInAlphabetics():
    """
    Write a program to print first name in alphabetical order
    :return:
    """
    List_names = []
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row
    print("#################### Names as it is  #######################")
    for i in range(2, row_count):
        index = 1
        cell = sheet.cell(row=i, column=2)
        name = cell.value
        List_names.append(name)
        print(name)

    print("###################  Names as per alphabets #################")
    List_names.sort()
    for item in List_names:
        print(item)


def getTotalCountBasedOnGender(gender):
    """
    WAP to print total number of employees based on the gender passed
    :param gender:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row
    col_count = sheet.max_column
    emp_count = 0
    try:
        for i in range(1, col_count + 1):
            cell_1 = sheet.cell(row=1, column=i)
            if cell_1.value == "gender":
                for k in range(1, row_count + 1):
                    cell_2 = sheet.cell(row=k, column=i)
                    if cell_2.value == gender:
                        emp_count = emp_count + 1
        print(f"Total {gender} employees count is:", emp_count)
    except Exception as e:
        print(e)


def employeesNameBasedOnGender(gender):
    """
    TO list the employess first and last name based on the gender
    :param gender:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row
    col_count = sheet.max_column
    fName_Column = 0
    lName_Column = 0
    print(f"\n List of {gender} employees:\n")
    for i in range(1, col_count + 1):
        cell1 = sheet.cell(row=1, column=i)
        if cell1.value == "first_name": fName_Column = i
        if cell1.value == "last_name": lName_Column = i
        if cell1.value == "gender":
            for k in range(1, row_count + 1):
                cell2 = sheet.cell(row=k, column=i)
                if cell2.value == gender:
                    f_name = sheet.cell(row=k, column=fName_Column)
                    l_name = sheet.cell(row=k, column=lName_Column)
                    print(f_name.value, '\t\t', l_name.value)


def printNamesAsperHiredMonth(month_name):
    """
    Print list of employees who are hired in a particular month
    :param month_name:
    :return:
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    row_count = sheet.max_row
    col_count = sheet.max_column
    fName_Column = 0
    lName_Column = 0
    print(f"\n List of employees hired in the month of {month_name}:\n")
    for i in range(1,col_count+1):
        cell_1 = sheet.cell(row=1,column=i)
        if cell_1.value == "first_name": fName_Column = i
        if cell_1.value == "last_name": lName_Column = i
        if cell_1.value == "hire_date":
            for k in range(2,row_count+1):
                cell_2 = sheet.cell(row=k, column=i)
                datetime_object = datetime.datetime.strptime(str(cell_2.value), "%Y-%m-%d %H:%M:%S")
                get_month_name = datetime_object.strftime("%b")
                if re.match(get_month_name,month_name,re.IGNORECASE):
                    f_name = sheet.cell(row=k, column=fName_Column)
                    l_name = sheet.cell(row=k, column=lName_Column)
                    print(f_name.value, l_name.value)




#getAllDuplicateRecords("Alexander")
#getValueInAlphabetics()
#getTotalCountBasedOnGender("Female")
#employeesNameBasedOnGender("Male")
#printNamesAsperHiredMonth("August")

